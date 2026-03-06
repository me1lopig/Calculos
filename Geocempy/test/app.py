from __future__ import annotations

import io
import logging
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import ezdxf
import streamlit as st
from ezdxf.enums import TextEntityAlignment
from openpyxl import load_workbook

RED = 1
GREEN = 3
BLUE = 5


@dataclass
class Model:
    version: str
    file_name: str
    uploaded_files: dict[str, bytes]
    profiles: list[str]
    draw_p: bool
    draw_s: bool
    read_ugeo: bool


@dataclass
class Penetro:
    name: str
    max_depth: float
    depths: list[float]
    beats: list[float]


@dataclass
class Estrato:
    depth_start: float
    depth_end: float
    cod: str
    des: str


@dataclass
class Ugeo:
    depth_start: float
    depth_end: float
    cod: str
    des: str


@dataclass
class Ensayo:
    depth_start: float
    geo: str
    spt: str
    mi: str
    uscs: str
    rcs: str
    roz: str
    coh: str


@dataclass
class Columna:
    name: str
    max_depth: float
    freatic: float
    lito: list[Estrato] = field(default_factory=list)
    ugeos: list[Ugeo] = field(default_factory=list)
    ensayos: list[Ensayo] = field(default_factory=list)


@dataclass
class Station:
    sta: str
    dist: float
    elev: float


@dataclass
class Profile:
    name: str
    max_height: float
    min_height: float
    max_length: float
    min_length: float
    stations: list[Station] = field(default_factory=list)


def setup_logger() -> Path:
    appdata = Path(os.getenv("APPDATA", "."))
    log_path = appdata / "Geocempy_3_0" / "Error.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=log_path,
        level=logging.ERROR,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    return log_path


def as_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_xlsx_rows(file_content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    col_count = max((len(r) for r in rows if r is not None), default=0)
    body = rows[1:]

    data: list[dict[str, Any]] = []
    for row in body:
        values = list(row or ())
        if len(values) < col_count:
            values.extend([None] * (col_count - len(values)))
        data.append({f"COL{i}": values[i] for i in range(col_count)})
    return data


def distinct_in_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def load_penetros(dprg: list[dict[str, Any]]) -> list[Penetro]:
    names = distinct_in_order([as_str(row.get("COL0")) for row in dprg])
    penetros: list[Penetro] = []
    for name in names:
        rows = [r for r in dprg if as_str(r.get("COL0")) == name]
        depths = [as_float(r.get("COL6")) for r in rows]
        beats = [as_float(r.get("COL7")) for r in rows]
        penetros.append(Penetro(name=name, max_depth=max(depths, default=0.0), depths=depths, beats=beats))
    return penetros


def load_columnas(
    bkfl: list[dict[str, Any]],
    lab: list[dict[str, Any]],
    ispt: list[dict[str, Any]],
    ugeo: list[dict[str, Any]],
    read_ugeo: bool,
) -> list[Columna]:
    names = distinct_in_order([as_str(row.get("COL0")) for row in bkfl])
    columnas: list[Columna] = []

    for name in names:
        lito = [
            Estrato(
                depth_start=as_float(r.get("COL2")),
                depth_end=as_float(r.get("COL3")),
                cod=as_str(r.get("COL5")),
                des=as_str(r.get("COL6")),
            )
            for r in bkfl
            if as_str(r.get("COL0")) == name
        ]

        ensayos = [
            Ensayo(
                depth_start=as_float(r.get("COL3")),
                geo=as_str(r.get("COL30")),
                spt=as_str(r.get("COL8")),
                mi=as_str(r.get("COL9")),
                uscs=as_str(r.get("COL21")),
                rcs=as_str(r.get("COL36")),
                roz=as_str(r.get("COL38")),
                coh=as_str(r.get("COL39")),
            )
            for r in lab
            if as_str(r.get("COL1")) == name
        ]

        ugeos: list[Ugeo] = []
        if read_ugeo:
            ugeos = [
                Ugeo(
                    depth_start=as_float(r.get("COL2")),
                    depth_end=as_float(r.get("COL3")),
                    cod=as_str(r.get("COL4")),
                    des=as_str(r.get("COL5")),
                )
                for r in ugeo
                if as_str(r.get("COL0")) == name
            ]

        nf_values = [
            as_float(r.get("COL6"))
            for r in ispt
            if as_str(r.get("COL0")) == name and r.get("COL6") is not None
        ]
        freatic = nf_values[0] if nf_values else -1.0

        columnas.append(
            Columna(
                name=name,
                max_depth=max((e.depth_end for e in lito), default=0.0),
                freatic=freatic,
                lito=lito,
                ugeos=ugeos,
                ensayos=ensayos,
            )
        )

    return columnas


def get_ugeo_names(ugeo: list[dict[str, Any]]) -> list[str]:
    return distinct_in_order([as_str(r.get("COL5")) for r in ugeo])


def load_profile(rows: list[dict[str, Any]], name: str) -> Profile:
    stations = [
        Station(
            sta=as_str(r.get("COL0")),
            dist=as_float(r.get("COL1")),
            elev=as_float(r.get("COL2")),
        )
        for r in rows
        if as_str(r.get("COL0"))
    ]

    return Profile(
        name=name,
        max_height=max((s.elev for s in stations), default=0.0),
        min_height=min((s.elev for s in stations), default=0.0),
        max_length=max((s.dist for s in stations), default=0.0),
        min_length=min((s.dist for s in stations), default=0.0),
        stations=stations,
    )


def new_block(doc: ezdxf.document.Drawing, name: str):
    if name in doc.blocks:
        doc.blocks.delete_block(name, safe=False)
    return doc.blocks.new(name)


def add_text(entity, text: str, pos: tuple[float, float], height: float, color: int | None = None, align=None):
    attrs: dict[str, Any] = {"height": height}
    if color is not None:
        attrs["color"] = color
    text_entity = entity.add_text(text, dxfattribs=attrs)
    if align is not None:
        text_entity.set_placement(pos, align=align)
    else:
        text_entity.dxf.insert = pos


def create_penetro(doc: ezdxf.document.Drawing, p: Penetro, max_x: float) -> str:
    scale = max_x / 100.0
    step = max_x / 4.0
    block = new_block(doc, p.name)

    add_text(block, p.name, (0, 2.5), 0.5, RED)

    block.add_line((0, 0), (max_x, 0), dxfattribs={"color": BLUE})
    block.add_line((0, 0), (0, 1), dxfattribs={"color": BLUE})
    block.add_line((step, 0), (step, 0.5), dxfattribs={"color": BLUE})
    block.add_line((2 * step, 0), (2 * step, 1.0), dxfattribs={"color": BLUE})
    block.add_line((3 * step, 0), (3 * step, 0.5), dxfattribs={"color": BLUE})
    block.add_line((4 * step, 0), (4 * step, 1.0), dxfattribs={"color": BLUE})

    add_text(block, "0", (0, 1.2), 0.3)
    add_text(block, "50", (2 * step, 1.2), 0.2)
    add_text(block, "R", (4 * step, 1.2), 0.3)

    block.add_line((0, 0), (0, -p.max_depth), dxfattribs={"color": BLUE})

    points = [(0.0, 0.0)]
    for depth, beat in zip(p.depths, p.beats):
        points.append((scale * beat, -depth))
    block.add_lwpolyline(points)
    return block.name


def ensayo_value(ensayo: Ensayo) -> str:
    if ensayo.geo == "SPT":
        return ensayo.spt
    if ensayo.geo == "MI":
        return ensayo.mi
    return "#"


def draw_column_base(block, name: str, litos: list[Estrato], ensayos: list[Ensayo], freatic: float):
    add_text(block, name, (0, 2.5), 0.5, RED, TextEntityAlignment.MIDDLE_LEFT)

    for e in litos:
        block.add_line((0, -e.depth_start), (-0.1, -e.depth_start), dxfattribs={"color": BLUE})
        block.add_line((-0.1, -e.depth_start), (-0.1, -e.depth_end), dxfattribs={"color": BLUE})
        block.add_line((-0.1, -e.depth_end), (0.1, -e.depth_end), dxfattribs={"color": BLUE})
        block.add_line((0.1, -e.depth_end), (0.1, -e.depth_start), dxfattribs={"color": BLUE})
        block.add_line((0.1, -e.depth_start), (0, -e.depth_start), dxfattribs={"color": BLUE})

        depth_med = 0.5 * (e.depth_start + e.depth_end)
        block.add_line((0.1, -e.depth_start), (0.6, -e.depth_start))
        add_text(block, e.des, (0.7, -depth_med), 0.2, align=TextEntityAlignment.MIDDLE_LEFT)

    for ensayo in ensayos:
        value = ensayo_value(ensayo)
        label = f"{ensayo.geo} {value}; P.I. {ensayo.depth_start}"
        add_text(block, label, (0.7, -ensayo.depth_start), 0.2, align=TextEntityAlignment.MIDDLE_LEFT)
        block.add_line((0.2, -ensayo.depth_start), (0.5, -ensayo.depth_start))

    if freatic != -1:
        block.add_lwpolyline([(-0.25, -freatic), (0.25, -freatic)], dxfattribs={"const_width": 0.25, "color": BLUE})
        add_text(block, f"N.F. > {freatic}", (-0.35, -freatic), 0.2, BLUE, TextEntityAlignment.MIDDLE_RIGHT)


def create_columna(doc: ezdxf.document.Drawing, c: Columna) -> str:
    block = new_block(doc, c.name)
    draw_column_base(block, c.name, c.lito, c.ensayos, c.freatic)
    return block.name


def create_ugeo(doc: ezdxf.document.Drawing, c: Columna) -> str:
    block = new_block(doc, f"{c.name}-UG")
    ugeo_as_lito = [Estrato(u.depth_start, u.depth_end, u.cod, u.des) for u in c.ugeos]
    draw_column_base(block, c.name, ugeo_as_lito, c.ensayos, c.freatic)
    return block.name


def create_profile(doc: ezdxf.document.Drawing, prof: Profile, ugeo_list: list[str], read_ugeo: bool) -> str:
    ymax = round(prof.max_height)
    ymin = round(prof.min_height)
    x_length = prof.max_length - prof.min_length
    ln = 50
    y_length = ln + (ymax - ymin) + 2

    prof_name = f"{prof.name}-UG" if read_ugeo else prof.name
    block = new_block(doc, prof_name)

    add_text(block, prof.name, (0, y_length + 2.5), 1.0, BLUE)

    block.add_line((0, -1), (x_length, -1))
    x0 = 0.0
    while x0 <= x_length:
        block.add_line((x0, -1), (x0, -2))
        if x0 % 25 == 0:
            block.add_line((x0, 0), (x0, y_length))
            add_text(block, f"{x0:g}", (x0, -2.3), 0.2, align=TextEntityAlignment.TOP_CENTER)
        x0 += 1

    y0 = 0.0
    z0 = ymin - ln
    block.add_line((-1, 0), (-1, y_length))
    while y0 <= y_length:
        block.add_line((-1, y0), (-2, y0))
        if z0 % 5 == 0:
            block.add_line((0, y0), (x_length, y0))
            add_text(block, f"{z0:g}", (-2.3, y0), 0.2, align=TextEntityAlignment.MIDDLE_RIGHT)
        y0 += 1
        z0 += 1

    terrain = [(s.dist, s.elev - ymin + ln) for s in prof.stations]
    if terrain:
        block.add_lwpolyline(terrain, dxfattribs={"color": GREEN})

    for sta in prof.stations:
        ref_name = f"{sta.sta}-UG" if read_ugeo and sta.sta in ugeo_list else sta.sta
        if ref_name in doc.blocks:
            block.add_blockref(ref_name, (sta.dist, sta.elev - ymin + ln))

    return block.name


def create_leyend(doc: ezdxf.document.Drawing, ugeo_list: list[str], read_ugeo: bool) -> str:
    block = new_block(doc, "Legend")

    add_text(block, "LEYENDA", (0, 0), 0.8, align=TextEntityAlignment.BOTTOM_LEFT)

    step = -0.5
    if read_ugeo:
        for ugeo in ugeo_list:
            vertices = [(0, step), (0, -1.0 + step), (1.5, -1.0 + step), (1.5, step), (0, step)]
            block.add_lwpolyline(vertices)
            add_text(block, ugeo, (1.8, -0.7 + step), 0.4, align=TextEntityAlignment.MIDDLE_LEFT)
            step -= 1.5

    add_text(block, "NIVEL FREATICO > Profundidad", (0, -1 + step), 0.4, BLUE, TextEntityAlignment.BOTTOM_LEFT)
    block.add_line((12, -0.5 + step), (13.5, -0.5 + step), dxfattribs={"color": BLUE})

    add_text(block, "CH/250,0/22,2/62,8", (0, -3 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "SPT 14; P.I. 3.6", (0, -4 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "MI 86; P.I. 3.0", (0, -5 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "TP #; P.I. 8.7", (0, -6 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)

    add_text(block, "Clasificacion(U.S.C.S.)/CS(KPa)/Fi(o)/C(KPa)", (8.5, -3 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "Ensayo SPT; Golpeo N30; Profundidad Inicial(m)", (8.5, -4 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "Muestra inalterada; Golpeo N30; Profundidad Inicial(m)", (8.5, -5 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "Testigo parafinado; Profundidad Inicial(m)", (8.5, -6 + step), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)

    step2 = 5 / 4
    x0 = 0
    y0 = -8 + step
    add_text(block, "Escala grafica N20 DPSH", (8.5, y0), 0.5, align=TextEntityAlignment.BOTTOM_LEFT)
    block.add_line((x0, y0), (x0 + 5, y0), dxfattribs={"color": BLUE})
    block.add_line((x0, y0), (x0, y0 + 1), dxfattribs={"color": BLUE})
    block.add_line((x0 + step2, y0), (x0 + step2, y0 + 0.5), dxfattribs={"color": BLUE})
    block.add_line((x0 + 2 * step2, y0), (x0 + 2 * step2, y0 + 1), dxfattribs={"color": BLUE})
    block.add_line((x0 + 3 * step2, y0), (x0 + 3 * step2, y0 + 0.5), dxfattribs={"color": BLUE})
    block.add_line((x0 + 4 * step2, y0), (x0 + 4 * step2, y0 + 1), dxfattribs={"color": BLUE})
    add_text(block, "0", (x0, y0 + 1.2), 0.3, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "50", (x0 + 2 * step2, y0 + 1.2), 0.3, align=TextEntityAlignment.BOTTOM_LEFT)
    add_text(block, "R", (x0 + 4 * step2, y0 + 1.2), 0.3, align=TextEntityAlignment.BOTTOM_LEFT)

    add_text(block, "Sondeo a rotacion", (40, -1), 0.5, align=TextEntityAlignment.BOTTOM_CENTER)
    add_text(block, "S-01", (40, -2), 0.5, RED, TextEntityAlignment.BOTTOM_CENTER)
    block.add_line((39.9, -3), (39.9, -9), dxfattribs={"color": BLUE})
    block.add_line((39.9, -9), (40.1, -9), dxfattribs={"color": BLUE})
    block.add_line((40.1, -9), (40.1, -3), dxfattribs={"color": BLUE})
    block.add_line((40.1, -3), (39.9, -3), dxfattribs={"color": BLUE})
    add_text(block, "Clasificacion/CS/Fi/C.", (39.5, -5), 0.3, align=TextEntityAlignment.MIDDLE_RIGHT)
    add_text(block, "Ensayo N30; Prof.Inicial.", (40.5, -5), 0.3, align=TextEntityAlignment.MIDDLE_LEFT)

    add_text(block, "Penetracion dinamica (DPSH)", (40, -11), 0.5, align=TextEntityAlignment.BOTTOM_CENTER)
    add_text(block, "P-01", (40, -12), 0.5, RED, TextEntityAlignment.BOTTOM_CENTER)
    block.add_line((40, -13), (40, -16.5), dxfattribs={"color": BLUE})

    return block.name


def get_uploaded_file(uploaded_files: dict[str, bytes], expected_name: str) -> bytes:
    expected_lower = expected_name.lower()
    for name, content in uploaded_files.items():
        if name.lower() == expected_lower:
            return content
    raise FileNotFoundError(f"No se encontro el archivo requerido: {expected_name}")


def generate_dxf(model: Model) -> bytes:
    bkfl = read_xlsx_rows(get_uploaded_file(model.uploaded_files, "BKFL.XLSX"))
    dprg = read_xlsx_rows(get_uploaded_file(model.uploaded_files, "DPRG.XLSX"))
    ispt = read_xlsx_rows(get_uploaded_file(model.uploaded_files, "ISPT.XLSX"))
    llab = read_xlsx_rows(get_uploaded_file(model.uploaded_files, "ListadoLab.XLSX"))
    ugeo = read_xlsx_rows(get_uploaded_file(model.uploaded_files, "UGEO.XLSX")) if model.read_ugeo else []

    profile_tables = [
        (Path(profile_name).stem, read_xlsx_rows(get_uploaded_file(model.uploaded_files, profile_name)))
        for profile_name in model.profiles
    ]

    penetros = load_penetros(dprg)
    columnas = load_columnas(bkfl, llab, ispt, ugeo, model.read_ugeo)
    profiles = [load_profile(rows, name) for name, rows in profile_tables]
    ugeo_names = get_ugeo_names(ugeo) if model.read_ugeo else []

    doc = ezdxf.new(dxfversion="R2010")
    msp = doc.modelspace()

    x = 0.0
    for p in penetros:
        block_name = create_penetro(doc, p, 5.0)
        if model.draw_p:
            msp.add_blockref(block_name, (x, 350))
            x += 20

    x = 0.0
    for c in columnas:
        block_name = create_columna(doc, c)
        if model.draw_s:
            msp.add_blockref(block_name, (x, 450))

        if model.read_ugeo:
            block_name2 = create_ugeo(doc, c)
            msp.add_blockref(block_name2, (x, 550))

        x += 20

    ugeo_list = [c.name for c in columnas if c.ugeos] if model.read_ugeo else []

    x = 0.0
    y = 0.0
    for prof in profiles:
        block_name = create_profile(doc, prof, ugeo_list, False)
        msp.add_blockref(block_name, (x, y))

        if model.read_ugeo:
            block_name2 = create_profile(doc, prof, ugeo_list, True)
            msp.add_blockref(block_name2, (x, y + 100))

        x += (prof.max_length - prof.min_length) + 50

    legend_name = create_leyend(doc, ugeo_names, model.read_ugeo)
    msp.add_blockref(legend_name, (0, -10))

    dxf_stream = io.StringIO()
    doc.write(dxf_stream)
    return dxf_stream.getvalue().encode("utf-8")


def list_profile_files(uploaded_files: dict[str, bytes]) -> list[str]:
    fixed = {"bkfl.xlsx", "dprg.xlsx", "ispt.xlsx", "listadolab.xlsx", "ugeo.xlsx"}
    return sorted(
        [
            name
            for name in uploaded_files.keys()
            if name.lower().endswith(".xlsx") and name.lower() not in fixed
        ]
    )


def main() -> None:
    log_file = setup_logger()

    st.set_page_config(page_title="Geocempy Streamlit", layout="wide")
    st.title("Geocempy 3.0 - Streamlit")

    uploaded_objects = st.file_uploader(
        "Subir archivos XLSX (base + perfiles)",
        type=["xlsx"],
        accept_multiple_files=True,
    )
    uploaded_files = {f.name: f.getvalue() for f in uploaded_objects}

    available_profiles = list_profile_files(uploaded_files)
    if available_profiles:
        profiles = st.multiselect("Perfiles", options=available_profiles)
    else:
        st.info(
            "No se detectaron archivos de perfil. "
            "Sube uno o mas XLSX adicionales (distintos de BKFL, DPRG, ISPT, ListadoLab y UGEO)."
        )
        profiles = []

    file_name = st.text_input("Fichero de salida", value="Default_DXF").strip()
    st.caption("Archivos base requeridos: BKFL.XLSX, DPRG.XLSX, ISPT.XLSX, ListadoLab.XLSX.")
    st.caption("Si activas unidades geotecnicas, tambien debes subir UGEO.XLSX.")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        version = st.selectbox("Version", ["R2010"], index=0)
    with c2:
        draw_p = st.checkbox("Representar DPSH", value=False)
    with c3:
        draw_s = st.checkbox("Representar Sondeos", value=False)
    with c4:
        read_ugeo = st.checkbox("Representar Unidades Geotecnicas", value=False)

    if st.button("Generar DXF", type="primary"):
        if not uploaded_files:
            st.error("Debes subir los archivos XLSX.")
            return
        if not file_name:
            st.error("El campo <<Fichero de salida>> esta vacio.")
            return

        required = ["BKFL.XLSX", "DPRG.XLSX", "ISPT.XLSX", "ListadoLab.XLSX"]
        if read_ugeo:
            required.append("UGEO.XLSX")

        missing = [
            name
            for name in required
            if all(uploaded.lower() != name.lower() for uploaded in uploaded_files.keys())
        ]
        if missing:
            st.error(f"Faltan archivos requeridos: {', '.join(missing)}")
            return
        if not available_profiles:
            st.error(
                "No hay perfiles disponibles. "
                "Sube al menos un archivo XLSX de perfil adicional y selecciona uno."
            )
            return
        if not profiles:
            st.error("El campo <<Perfiles>> esta vacio.")
            return

        model = Model(
            version=version,
            file_name=file_name,
            uploaded_files=uploaded_files,
            profiles=profiles,
            draw_p=draw_p,
            draw_s=draw_s,
            read_ugeo=read_ugeo,
        )

        try:
            output_bytes = generate_dxf(model)
            st.session_state["generated_dxf"] = output_bytes
            st.session_state["generated_dxf_name"] = f"{file_name}.dxf"
            st.success("El archivo DXF se ha generado correctamente.")
        except Exception as exc:
            logging.exception("DXF generation failed")
            st.error(f"Se ha producido un error. Revise el log: {log_file}")
            st.exception(exc)

    if "generated_dxf" in st.session_state and "generated_dxf_name" in st.session_state:
        st.download_button(
            label="Descargar DXF",
            data=st.session_state["generated_dxf"],
            file_name=st.session_state["generated_dxf_name"],
            mime="application/dxf",
        )


if __name__ == "__main__":
    main()
