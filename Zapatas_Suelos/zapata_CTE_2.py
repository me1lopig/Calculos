import math
from openpyxl import Workbook
import openpyxl.styles
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import os

def calcular_factores_capacidad(phi_deg):
    """
    Calcula factores de capacidad portante según Brinch-Hansen (1970).
    
    Parámetros:
        phi_deg (float): Ángulo de rozamiento interno en grados [0-45°]
    
    Retorna:
        tuple: (Nc, Nq, Ngamma) - Factores de capacidad portante
        - Nc: Factor para cohesión
        - Nq: Factor para sobrecarga
        - Ngamma: Factor para peso unitario del suelo
    
    Referencia: Brinch-Hansen, A. (1970). A revised and extended formula for bearing capacity
    """
    if not 0 <= phi_deg <= 45:
        raise ValueError(f"Ángulo φ debe estar entre 0° y 45°, obtuvo {phi_deg}°")
    
    phi_rad = math.radians(phi_deg)
    
    # Protección contra división por cero cuando φ=0
    if phi_deg == 0:
        Nc = 5.14  # Valor teórico para suelo puramente cohesivo
        Nq = 1.0   # No hay contribución de sobrecarga
        Ngamma = 0.0  # No hay contribución del peso del suelo
    else:
        Nq = math.exp(math.pi * math.tan(phi_rad)) * (math.tan(math.pi/4 + phi_rad/2))**2
        Nc = (Nq - 1) / math.tan(phi_rad)
        Ngamma = 2 * (Nq + 1) * math.tan(phi_rad)
    
    return Nc, Nq, Ngamma

def peso_especifico_sumergido(gamma_sat, gamma_w=9.81):
    """
    Calcula el peso específico sumergido (γ').
    
    Parámetros:
        gamma_sat (float): Peso específico saturado en kN/m³
        gamma_w (float): Peso específico del agua en kN/m³ (default=9.81)
    
    Retorna:
        float: Peso específico sumergido (γ' = γ_sat - γ_w)
    
    Notas:
        γ' es el peso efectivo bajo el nivel freático
    """
    return gamma_sat - gamma_w

def factores_forma(B, L):
    """
    Calcula factores de forma para cimentación rectangular.
    
    Parámetros:
        B (float): Ancho de la zapata en metros
        L (float): Largo de la zapata en metros
    
    Retorna:
        tuple: (sc, sq, sgamma) - Factores de forma
        - sc: Factor de forma para cohesión (rango [1, 1.3])
        - sq: Factor de forma para sobrecarga (rango [1, 1.3])
        - sgamma: Factor de forma para peso del suelo (rango [0.6, 1])
    
    Notas:
        - Para zapata cuadrada (B=L): sc=sq=1.2, sgamma=0.8
        - Fórmulas de Brinch-Hansen (1970)
    """
    if L == 0:
        return 1.0, 1.0, 0.6
    sc = 1 + 0.2 * (B / L)
    sq = 1 + 0.2 * (B / L)
    sgamma = 1 - 0.4 * (B / L)
    return sc, sq, sgamma

def factores_profundidad(D, B):
    """
    Calcula factores de profundidad de empotramiento.
    
    Parámetros:
        D (float): Profundidad de empotramiento en metros
        B (float): Ancho de la zapata en metros
    
    Retorna:
        tuple: (dc, dq, dgamma) - Factores de profundidad
        - dc: Factor de profundidad para cohesión
        - dq: Factor de profundidad para sobrecarga
        - dgamma: Factor de profundidad para peso del suelo (siempre 1.0)
    
    Notas:
        - Los factores aumentan con D/B (mayor profundidad → mayor capacidad)
        - Fórmulas de Brinch-Hansen (1970)
    """
    if B == 0:
        return 1.0, 1.0, 1.0
    dq = 1 + 0.2 * (D / B)
    dc = 1 + 0.2 * (D / B)
    dgamma = 1.0
    return dc, dq, dgamma

def factores_inclinacion(alpha_deg=0):
    """
    Calcula factores de inclinación de carga (Brinch-Hansen 1970).
    
    Parámetros:
        alpha_deg (float): Ángulo de inclinación de la carga en grados [0-90°]
    
    Retorna:
        tuple: (ic, iq, igamma) - Factores de inclinación
    
    Notas:
        - α = 0° → carga vertical pura (factores = 1)
        - α = 90° → carga horizontal pura (factores = 0)
    """
    if not 0 <= alpha_deg <= 90:
        raise ValueError(f"Ángulo α debe estar entre 0° y 90°, obtuvo {alpha_deg}°")
    
    ic = max(0, 1 - alpha_deg / 45)  # Lineal, rango [0, 1]
    iq = max(0, (1 - alpha_deg / 90)**2)  # Cuadrático, rango [0, 1]
    igamma = max(0, (1 - alpha_deg / 90)**2)  # Cuadrático, rango [0, 1]
    return ic, iq, igamma

def obtener_float(mensaje, min_val=None, max_val=None):
    """
    Solicita un valor float al usuario con validación robusta.
    
    Parámetros:
        mensaje (str): Texto a mostrar al usuario
        min_val (float): Valor mínimo permitido (inclusive)
        max_val (float): Valor máximo permitido (inclusive)
    
    Retorna:
        float: Valor validado ingresado por el usuario
    """
    while True:
        try:
            valor = float(input(mensaje))
            if min_val is not None and valor < min_val:
                print(f"  ⚠️  El valor debe ser >= {min_val}. Inténtalo de nuevo.")
                continue
            if max_val is not None and valor > max_val:
                print(f"  ⚠️  El valor debe ser <= {max_val}. Inténtalo de nuevo.")
                continue
            return valor
        except ValueError:
            print("  ❌ Entrada no válida. Introduce un número decimal (ej: 2.5)")

def preguntar_sobrecarga():
    """
    Solicita al usuario si desea considerar sobrecarga de tierras.
    
    Retorna:
        bool: True si desea incluir sobrecarga, False en caso contrario
    
    Notas:
        - Con sobrecarga: q = γ·D (presión total)
        - Sin sobrecarga: solo contribución de la cohesión y peso del suelo
    """
    while True:
        respuesta = input("\n¿Considerar sobrecarga de tierras? (s/n): ").strip().lower()
        if respuesta in ['s', 'n', 'si', 'no']:
            return respuesta in ['s', 'si']
        print("  ❌ Respuesta no válida. Introduce 's' (sí) o 'n' (no).")

def calcular_carga_admisible(B, L, D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg, considerar_sobrecarga):
    """
    Calcula la capacidad portante admisible según Brinch-Hansen (1970).
    
    Parámetros:
        B (m): Ancho de la zapata
        L (m): Largo de la zapata
        D (m): Profundidad de empotramiento
        c (kPa): Cohesión efectiva
        phi_deg (°): Ángulo de rozamiento interno [0-45]
        gamma (kN/m³): Peso específico natural
        gamma_sat (kN/m³): Peso específico saturado
        D_w (m): Profundidad del nivel freático desde superficie
        FS: Factor de seguridad
        alpha_deg (°): Inclinación de la carga [0-90]
        considerar_sobrecarga (bool): Incluir sobrecarga de tierras
    
    Retorna:
        tuple: (q_ult, q_net_ult, q_adm, carga_adm_total) en kPa y kN
    
    Fórmula:
        q_ult = c·Nc·sc·dc·ic + q·Nq·sq·dq·iq + 0.5·γ'·B·Nγ·sγ·dγ·iγ
        q_net_ult = q_ult - γ_eff·D (presión efectiva en la base)
        q_adm = q_net_ult / FS
    
    Notas:
        - γ_eff = γ_sat - γ_w si D_w <= D (base bajo freático)
        - γ_eff = γ si D_w > D (base sobre freático)
    """
    Nc, Nq, Ngamma = calcular_factores_capacidad(phi_deg)
    sc, sq, sgamma = factores_forma(B, L)
    dc, dq, dgamma = factores_profundidad(D, B)
    ic, iq, igamma = factores_inclinacion(alpha_deg)

    # Término de cohesión
    termino_c = c * Nc * sc * dc * ic

    # Término de sobrecarga: considerar nivel freático
    if considerar_sobrecarga:
        if D_w <= D:
            # Freático dentro del cimiento: estratificación
            gamma_seco = gamma
            gamma_sumergido = peso_especifico_sumergido(gamma_sat)
            # q = γ_seco·D_w + γ_sum·(D - D_w)
            q = gamma_seco * D_w + gamma_sumergido * (D - D_w)
            termino_q = q * Nq * sq * dq * iq
        else:
            # Freático bajo la base: usar peso seco
            termino_q = gamma * D * Nq * sq * dq * iq
    else:
        termino_q = 0

    # Termo de peso del suelo (debajo de la base)
    if D_w <= D:
        gamma_efectivo = peso_especifico_sumergido(gamma_sat)
    else:
        gamma_efectivo = gamma
    
    termino_gamma = 0.5 * gamma_efectivo * B * Ngamma * sgamma * dgamma * igamma
    
    # Capacidad portante última
    q_ult = termino_c + termino_q + termino_gamma
    
    # Presión efectiva en la base (para restar en capacidad neta)
    if D_w <= D:
        gamma_base = peso_especifico_sumergido(gamma_sat)
    else:
        gamma_base = gamma
    
    q_net_ult = q_ult - gamma_base * D
    q_adm = q_net_ult / FS
    carga_adm_total = q_adm * B * L

    return q_ult, q_net_ult, q_adm, carga_adm_total

def generar_combinaciones_B_L(B_inicio, B_fin, B_paso, L_inicio, L_fin, L_paso):
    """
    Genera combinaciones de dimensiones (B, L) evitando errores de punto flotante.
    
    Parámetros:
        B_inicio, B_fin, B_paso (float): Rango e incremento de ancho
        L_inicio, L_fin, L_paso (float): Rango e incremento de largo
    
    Retorna:
        list: Lista de tuplas (B, L) donde B <= L, ordenadas
    
    Notas:
        - Usa conteo entero para evitar acumulación de errores flotantes
        - Aplica redondeo a 2 decimales al final
        - Garantiza que todos los valores respetan los límites
    """
    # Usar índices enteros y truncamiento para garantizar límites
    tolerancia = 1e-9
    valores_B = []
    valor = B_inicio
    while valor <= B_fin + tolerancia:
        valores_B.append(round(valor, 3))
        valor += B_paso
    
    valores_L = []
    valor = L_inicio
    while valor <= L_fin + tolerancia:
        valores_L.append(round(valor, 3))
        valor += L_paso
    
    # Eliminar duplicados por redondeo y garantizar límites
    valores_B = sorted(set([min(v, B_fin) for v in valores_B]))
    valores_L = sorted(set([min(v, L_fin) for v in valores_L]))
    
    # Filtrar: B <= L y redondear a 2 decimales
    combinaciones = [(round(B, 2), round(L, 2)) for B in valores_B for L in valores_L if B <= L]
    return combinaciones

def exportar_a_excel(resultados, nombre_archivo="resultados_cimentacion.xlsx", parametros=None):
    """
    Exporta los resultados a un archivo Excel con formato mejorado.
    
    Parámetros:
        resultados (list): Lista de [B, L, q_ult, q_net_ult, q_adm, carga_adm_total]
        nombre_archivo (str): Nombre del archivo Excel a generar
        parametros (dict): Diccionario con parámetros de entrada para documentar
    
    Genera:
        - Archivo Excel con datos formateados
        - Encabezado con parámetros de entrada (opcional)
        - Tabla de resultados con ancho automático
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    # Si hay parámetros, agregar encabezado informativo
    if parametros:
        ws['A1'] = "CÁLCULO DE CARGA ADMISIBLE - CTE-DB-SE-C"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "PARÁMETROS DE ENTRADA"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=11)
        
        row += 1
        param_labels = [
            ("Profundidad empotramiento (D):", "D"),
            ("Cohesión efectiva (c):", "c"),
            ("Ángulo rozamiento (φ):", "phi"),
            ("Peso específico natural (γ):", "gamma"),
            ("Peso específico saturado (γ_sat):", "gamma_sat"),
            ("Profundidad nivel freático (D_w):", "D_w"),
            ("Factor de seguridad (FS):", "FS"),
            ("Inclinación carga (α):", "alpha"),
            ("Considerar sobrecarga:", "sobrecarga"),
        ]
        
        for label, key in param_labels:
            if key in parametros:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = parametros[key]
                row += 1
        
        row += 1
    else:
        row = 1
    
    # Encabezado de tabla
    headers = ["B (m)", "L (m)", "q_ult (kPa)", "q_net_ult (kPa)", "q_adm (kPa)", "Carga admisible (kN)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Datos
    for i, resultado in enumerate(resultados, row + 1):
        for j, valor in enumerate(resultado, 1):
            cell = ws.cell(row=i, column=j, value=valor)
            if j > 2:  # Columnas numéricas con decimales
                cell.number_format = '0.00'
    
    # Ancho automático
    for col in ws.columns:
        max_length = 15
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    wb.save(nombre_archivo)
    print(f"\n✓ Resultados exportados a: {nombre_archivo}")


def graficar_lineas_multiples(resultados, parametros=None, directorio=None):
    """
    Genera gráfico de líneas múltiples: q_adm vs B para cada valor de L.
    
    Parámetros:
        resultados (list): Lista de [B, L, q_ult, q_net_ult, q_adm, carga_adm_total]
        parametros (dict): Diccionario con parámetros de entrada (opcional)
        directorio (str): Directorio donde guardar la imagen (opcional)
    
    Genera:
        - Gráfico con líneas para cada L
        - Archivo PNG de alta resolución
        - Título descriptivo con parámetros
    
    Retorna:
        str: Ruta del archivo guardado
    """
    if not resultados:
        print("  ❌ No hay resultados para graficar")
        return None
    
    # Crear figura
    fig, ax = plt.subplots(figsize=(12, 7))
    
    # Agrupar resultados por valor de L
    L_values = sorted(set([r[1] for r in resultados]))
    
    # Colores y estilos
    colores = plt.cm.tab10(np.linspace(0, 1, len(L_values)))
    
    # Plotear líneas para cada L
    for idx, L_value in enumerate(L_values):
        B_vals = [r[0] for r in resultados if r[1] == L_value]
        q_adm_vals = [r[4] for r in resultados if r[1] == L_value]
        
        # Ordenar por B
        pares = sorted(zip(B_vals, q_adm_vals))
        B_vals_sorted = [p[0] for p in pares]
        q_adm_sorted = [p[1] for p in pares]
        
        ax.plot(B_vals_sorted, q_adm_sorted, 
               marker='o', linewidth=2.5, markersize=7,
               label=f'L = {L_value:.2f} m', color=colores[idx], alpha=0.8)
    
    # Configurar etiquetas y título
    ax.set_xlabel('Ancho B (m)', fontsize=13, fontweight='bold')
    ax.set_ylabel('Capacidad Admisible q_adm (kPa)', fontsize=13, fontweight='bold')
    
    # Título con parámetros
    if parametros:
        titulo = f"Capacidad Portante Admisible vs Dimensiones\n"
        titulo += f"φ={parametros.get('phi', '?')}°, c={parametros.get('c', '?')} kPa, "
        titulo += f"FS={parametros.get('FS', '?')}, D={parametros.get('D', '?')} m"
    else:
        titulo = "Capacidad Portante Admisible vs Dimensiones"
    
    ax.set_title(titulo, fontsize=14, fontweight='bold', pad=20)
    
    # Grid y leyenda
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(fontsize=11, loc='best', framealpha=0.95, edgecolor='black')
    
    # Mejorar aspecto
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_facecolor('#f8f9fa')
    fig.patch.set_facecolor('white')
    
    # Ajustar layout
    plt.tight_layout()
    
    # Guardar archivo
    if directorio is None:
        directorio = "."
    
    # Crear nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = os.path.join(directorio, f"q_adm_lineas_{timestamp}.png")
    
    try:
        plt.savefig(nombre_archivo, dpi=300, bbox_inches='tight', facecolor='white')
        print(f"\n✓ Gráfico guardado en: {nombre_archivo}")
    except Exception as e:
        print(f"  ⚠️  Error al guardar gráfico: {e}")
        nombre_archivo = None
    
    plt.show()
    
    return nombre_archivo


def validar_coherencia_datos(D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg):
    """
    Valida coherencia física de los datos de entrada.
    
    Parámetros:
        Ver función main()
    
    Retorna:
        tuple: (valido, mensaje) donde valido es bool y mensaje describe cualquier problema
    
    Levanta:
        ValueError: Si hay incoherencias no recuperables
    """
    errores = []
    avisos = []
    
    # Validaciones críticas
    if gamma_sat < gamma:
        errores.append(f"⚠️  γ_saturado ({gamma_sat}) < γ_natural ({gamma}) - Físicamente imposible")
    
    if D_w > D:
        avisos.append(f"ℹ️  Nivel freático bajo la base (D_w={D_w}m > D={D}m) - Se asume D_w=D")
    
    if D_w < 0:
        errores.append(f"⚠️  Nivel freático negativo (D_w={D_w}m) - Inválido")
    
    if c < 0:
        errores.append(f"⚠️  Cohesión negativa (c={c}kPa) - Inválido")
    
    if phi_deg < 0 or phi_deg > 45:
        errores.append(f"⚠️  Ángulo φ fuera de rango [{phi_deg}°] - Debe estar en [0°, 45°]")
    
    if gamma < 10 or gamma > 30:
        avisos.append(f"ℹ️  γ_natural = {gamma} kN/m³ - Fuera del rango típico [10-30]")
    
    if gamma_sat < 15 or gamma_sat > 25:
        avisos.append(f"ℹ️  γ_saturado = {gamma_sat} kN/m³ - Fuera del rango típico [15-25]")
    
    if FS < 1 or FS > 10:
        avisos.append(f"ℹ️  Factor seguridad FS = {FS} - Rango típico [2-4]")
    
    # Mostrar errores y avisos
    if errores:
        for error in errores:
            print(f"  {error}")
        return False, "Datos incoherentes"
    
    if avisos:
        for aviso in avisos:
            print(f"  {aviso}")
    
    return True, "Datos validados correctamente"


def main():
    """
    Función principal: orquestador del cálculo de capacidad portante.
    
    Flujo:
        1. Solicitar parámetros del terreno
        2. Validar coherencia de datos
        3. Solicitar rangos de dimensiones
        4. Confirmar parámetros antes de calcular
        5. Generar combinaciones B×L
        6. Calcular para cada combinación
        7. Mostrar resultados en tabla
        8. Exportar a Excel (opcional)
    """
    print("\n" + "="*80)
    print("   CÁLCULO DE CARGA ADMISIBLE PARA CIMENTACIONES SUPERFICIALES (CTE-DB-SE-C)")
    print("="*80)

    # ====== ENTRADA DE DATOS ======
    print("\n📋 PARÁMETROS DEL TERRENO:")
    print("-" * 80)
    
    D = obtener_float("  Profundidad de empotramiento (D, metros): ", 0.1)
    c = obtener_float("  Cohesión efectiva (c, kPa): ", 0)
    phi_deg = obtener_float("  Ángulo de rozamiento interno (φ, grados): ", 0, 45)
    gamma = obtener_float("  Peso específico natural (γ, kN/m³): ", 10, 30)
    gamma_sat = obtener_float("  Peso específico saturado (γ_sat, kN/m³): ", gamma, 30)
    D_w = obtener_float("  Profundidad del nivel freático (D_w, metros): ", 0)
    FS = obtener_float("  Factor de seguridad (FS, rango [1-10], default=3.0): ", 1, 10)
    alpha_deg = obtener_float("  Inclinación de la carga (α, grados, rango [0-90]): ", 0, 90)
    
    # Establecer valores por defecto si el usuario ingresa valores en los límites
    if FS < 1.5:
        FS = 3.0
    if alpha_deg > 85:
        alpha_deg = 0.0
    
    # ====== VALIDAR COHERENCIA ======
    print("\n✓ Validando coherencia de datos...")
    valido, msg = validar_coherencia_datos(D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg)
    if not valido:
        print(f"  ❌ {msg}")
        print("  Por favor, corrige los valores.")
        return
    print(f"  ✓ {msg}")
    
    # ====== SOBRECARGA ======
    considerar_sobrecarga = preguntar_sobrecarga()
    
    # ====== RANGOS DE DIMENSIONES ======
    print("\n📏 DIMENSIONES DE LA ZAPATA:")
    print("-" * 80)
    B_inicio = obtener_float("  Valor inicial de B (ancho, metros): ", 0.1)
    B_fin = obtener_float("  Valor final de B (metros): ", B_inicio)
    B_paso = obtener_float("  Incremento de B (metros): ", 0.01)

    L_inicio = obtener_float("  Valor inicial de L (largo, metros): ", B_inicio)
    L_fin = obtener_float("  Valor final de L (metros): ", L_inicio)
    L_paso = obtener_float("  Incremento de L (metros): ", 0.01)

    # Validar que los pasos sean razonables
    if B_paso > (B_fin - B_inicio) * 0.5:
        print(f"  ⚠️  B_paso es muy grande. Se ajusta a {(B_fin - B_inicio) * 0.1:.3f}m")
        B_paso = (B_fin - B_inicio) * 0.1
    
    if L_paso > (L_fin - L_inicio) * 0.5:
        print(f"  ⚠️  L_paso es muy grande. Se ajusta a {(L_fin - L_inicio) * 0.1:.3f}m")
        L_paso = (L_fin - L_inicio) * 0.1

    # ====== CONFIRMACIÓN ======
    print("\n📝 RESUMEN DE PARÁMETROS:")
    print("-" * 80)
    print(f"  Terreno: φ={phi_deg}°, c={c}kPa, γ={gamma}kN/m³, γ_sat={gamma_sat}kN/m³")
    print(f"  Cimentación: D={D}m, Freático: D_w={D_w}m")
    print(f"  Factor seguridad: FS={FS}, Inclinación: α={alpha_deg}°")
    print(f"  Sobrecarga: {'SÍ' if considerar_sobrecarga else 'NO'}")
    print(f"  Rango B: [{B_inicio}, {B_fin}] m, paso={B_paso}m")
    print(f"  Rango L: [{L_inicio}, {L_fin}] m, paso={L_paso}m")
    
    confirmacion = input("\n¿Proceeder con el cálculo? (s/n): ").strip().lower()
    if confirmacion not in ['s', 'si']:
        print("  Operación cancelada.")
        return

    # ====== GENERAR COMBINACIONES ======
    print("\n⚙️  Procesando...")
    combinaciones = generar_combinaciones_B_L(B_inicio, B_fin, B_paso, L_inicio, L_fin, L_paso)

    if not combinaciones:
        print(f"  ❌ No hay combinaciones válidas (B <= L)")
        print(f"  Ajusta los intervalos: L_inicio >= B_fin")
        return

    print(f"  ✓ Generadas {len(combinaciones)} combinaciones")

    # ====== CALCULAR ======
    resultados = []
    for B, L in combinaciones:
        try:
            q_ult, q_net_ult, q_adm, carga_adm_total = calcular_carga_admisible(
                B, L, D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg, considerar_sobrecarga
            )
            resultados.append([B, L, q_ult, q_net_ult, q_adm, carga_adm_total])
        except Exception as e:
            print(f"  ⚠️  Error en combinación B={B}m, L={L}m: {e}")
            continue

    if not resultados:
        print("  ❌ No se generaron resultados válidos.")
        return

    # ====== MOSTRAR RESULTADOS ======
    print("\n" + "="*80)
    print("📊 RESULTADOS DEL CÁLCULO")
    print("="*80)
    print(f"{'B (m)':<8} {'L (m)':<8} {'q_ult (kPa)':<14} {'q_net_ult (kPa)':<17} {'q_adm (kPa)':<14} {'Carga adm (kN)':<15}")
    print("-" * 80)
    for resultado in resultados:
        print(f"{resultado[0]:<8.2f} {resultado[1]:<8.2f} {resultado[2]:<14.2f} {resultado[3]:<17.2f} {resultado[4]:<14.2f} {resultado[5]:<15.2f}")

    # ====== EXPORTACIÓN ======
    exportar = input("\n¿Deseas exportar los resultados a Excel? (s/n): ").strip().lower()
    if exportar in ['s', 'si']:
        nombre_archivo = input("  Nombre del archivo (sin extensión, default='resultados'): ").strip() or "resultados_cimentacion"
        if not nombre_archivo.endswith('.xlsx'):
            nombre_archivo += '.xlsx'
        
        parametros = {
            'D': D,
            'c': c,
            'phi': phi_deg,
            'gamma': gamma,
            'gamma_sat': gamma_sat,
            'D_w': D_w,
            'FS': FS,
            'alpha': alpha_deg,
            'sobrecarga': 'Sí' if considerar_sobrecarga else 'No'
        }
        
        try:
            exportar_a_excel(resultados, nombre_archivo, parametros)
        except Exception as e:
            print(f"  ❌ Error al exportar: {e}")
    
    print("\n✓ Cálculo completado.\n")

# --- Ejecutar el programa ---
if __name__ == "__main__":
    try:
        main()
    except ImportError as e:
        print(f"\n❌ Error de librería: {e}")
        print("   Instala openpyxl con: pip install openpyxl")
    except KeyboardInterrupt:
        print("\n\n⚠️  Operación cancelada por el usuario (Ctrl+C).")
    except ZeroDivisionError as e:
        print(f"\n❌ Error matemático (división por cero): {e}")
        print("   Verifica que los valores de entrada sean válidos.")
    except ValueError as e:
        print(f"\n❌ Error de validación: {e}")
    except Exception as e:
        print(f"\n❌ Error inesperado ({type(e).__name__}): {e}")
        import traceback
        traceback.print_exc()