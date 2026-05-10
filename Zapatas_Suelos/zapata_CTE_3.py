import math
from openpyxl import Workbook
import openpyxl.styles
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import os

def calcular_factores_capacidad(phi_deg):
    if not 0 <= phi_deg <= 45:
        raise ValueError(f"Ángulo φ debe estar entre 0° y 45°, obtuvo {phi_deg}°")
    phi_rad = math.radians(phi_deg)
    if phi_deg == 0:
        Nc, Nq, Ngamma = 5.14, 1.0, 0.0
    else:
        Nq = math.exp(math.pi * math.tan(phi_rad)) * (math.tan(math.pi/4 + phi_rad/2))**2
        Nc = (Nq - 1) / math.tan(phi_rad)
        Ngamma = 2 * (Nq + 1) * math.tan(phi_rad)
    return Nc, Nq, Ngamma

def peso_especifico_sumergido(gamma_sat, gamma_w=9.81):
    return gamma_sat - gamma_w

def factores_forma(B, L):
    if L == 0: return 1.0, 1.0, 0.6
    sc = 1 + 0.2 * (B / L)
    sq = 1 + 0.2 * (B / L)
    sgamma = 1 - 0.4 * (B / L)
    return sc, sq, sgamma

def factores_profundidad(D, B):
    if B == 0: return 1.0, 1.0, 1.0
    dq = 1 + 0.2 * (D / B)
    dc = 1 + 0.2 * (D / B)
    return dc, dq, 1.0

def factores_inclinacion(alpha_deg=0):
    if not 0 <= alpha_deg <= 90:
        raise ValueError(f"Ángulo α debe estar entre 0° y 90°, obtuvo {alpha_deg}°")
    ic = max(0, 1 - alpha_deg / 45)
    iq = max(0, (1 - alpha_deg / 90)**2)
    igamma = max(0, (1 - alpha_deg / 90)**2)
    return ic, iq, igamma

def obtener_float(mensaje, min_val=None, max_val=None):
    while True:
        try:
            valor = float(input(mensaje))
            if min_val is not None and valor < min_val:
                print(f"  ⚠️ El valor debe ser >= {min_val}. Inténtalo de nuevo.")
                continue
            if max_val is not None and valor > max_val:
                print(f"  ⚠️ El valor debe ser <= {max_val}. Inténtalo de nuevo.")
                continue
            return valor
        except ValueError:
            print("  ❌ Entrada no válida. Introduce un número decimal.")

def preguntar_sobrecarga():
    while True:
        respuesta = input("\n¿Considerar sobrecarga de tierras? (s/n): ").strip().lower()
        if respuesta in ['s', 'n', 'si', 'no']:
            return respuesta in ['s', 'si']

def calcular_carga_admisible(B, L, D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg, considerar_sobrecarga):
    Nc, Nq, Ngamma = calcular_factores_capacidad(phi_deg)
    sc, sq, sgamma = factores_forma(B, L)
    dc, dq, dgamma = factores_profundidad(D, B)
    ic, iq, igamma = factores_inclinacion(alpha_deg)

    termino_c = c * Nc * sc * dc * ic

    if considerar_sobrecarga:
        if D_w <= D:
            q = gamma * D_w + peso_especifico_sumergido(gamma_sat) * (D - D_w)
        else:
            q = gamma * D
        termino_q = q * Nq * sq * dq * iq
    else:
        termino_q = 0

    gamma_efectivo = peso_especifico_sumergido(gamma_sat) if D_w <= D else gamma
    termino_gamma = 0.5 * gamma_efectivo * B * Ngamma * sgamma * dgamma * igamma
    
    q_ult = termino_c + termino_q + termino_gamma
    
    gamma_base = peso_especifico_sumergido(gamma_sat) if D_w <= D else gamma
    q_net_ult = q_ult - gamma_base * D
    q_adm = q_net_ult / FS
    carga_adm_total = q_adm * B * L

    return q_ult, q_net_ult, q_adm, carga_adm_total

def generar_combinaciones_B_L(B_inicio, B_fin, B_paso, L_inicio, L_fin, L_paso):
    tolerancia = 1e-9
    valores_B = []
    valor = B_inicio
    while valor <= B_fin + tolerancia:
        valores_B.append(round(valor, 3))
        valor += B_paso
    
    valores_L = []
    valor = L_inicio
    while valor <= L_fin + tolerancia:
        valores_L.append(round(valor, 3))
        valor += L_paso
    
    valores_B = sorted(set([min(v, B_fin) for v in valores_B]))
    valores_L = sorted(set([min(v, L_fin) for v in valores_L]))
    
    return [(round(B, 2), round(L, 2)) for B in valores_B for L in valores_L if B <= L]

def exportar_a_excel(resultados, ruta_archivo, parametros=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    if parametros:
        ws['A1'] = "CÁLCULO DE CARGA ADMISIBLE - CTE-DB-SE-C"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        row = 3
        ws[f'A{row}'] = "PARÁMETROS DE ENTRADA"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=11)
        row += 1
        for key, val in parametros.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = val
            row += 1
        row += 1
    else:
        row = 1
    
    headers = ["B (m)", "L (m)", "q_ult (kPa)", "q_net_ult (kPa)", "q_adm (kPa)", "Carga admisible (kN)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for i, resultado in enumerate(resultados, row + 1):
        for j, valor in enumerate(resultado, 1):
            cell = ws.cell(row=i, column=j, value=valor)
            if j > 2: cell.number_format = '0.00'
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    wb.save(ruta_archivo)

def generar_grafico_especifico(resultados, parametros, ruta_archivo):
    """
    Genera un único gráfico profesional de q_adm vs B, 
    con series de colores para cada valor de L.
    """
    if not resultados:
        return

    # Extraer datos
    B_vals = np.array([r[0] for r in resultados])
    L_vals = np.array([r[1] for r in resultados])
    q_adm_vals = np.array([r[4] for r in resultados])

    # Crear la figura
    plt.figure(figsize=(10, 6))
    
    # Obtener valores únicos de L para las series
    L_unicos = sorted(list(set(L_vals)))
    
    # Usar una paleta de colores profesional
    cmap = plt.get_cmap('viridis', len(L_unicos))

    for i, L_u in enumerate(L_unicos):
        # Filtrar datos para el valor actual de L
        mask = L_vals == L_u
        # Ordenar por B para asegurar que la línea sea continua
        indices = np.argsort(B_vals[mask])
        
        plt.plot(B_vals[mask][indices], q_adm_vals[mask][indices], 
                 marker='o', 
                 linestyle='-', 
                 linewidth=2, 
                 markersize=5,
                 label=f'L = {L_u:.2f} m',
                 color=cmap(i))

    # Configuración de ejes y etiquetas
    plt.title(f"Capacidad Portante Admisible (q_adm)\nParámetros: φ={parametros['phi']}°, c={parametros['c']} kPa, D={parametros['D']} m", 
              fontsize=12, fontweight='bold', pad=15)
    
    plt.xlabel("Ancho de la zapata B (m)", fontsize=11)
    plt.ylabel("Tensión Admisible q_adm (kPa)", fontsize=11)
    
    # Añadir cuadrícula (grid) técnica
    plt.grid(True, which='both', linestyle='--', alpha=0.5)
    
    # Leyenda fuera del gráfico para no tapar líneas
    plt.legend(title="Largo de zapata", bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)

    # Ajustar diseño
    plt.tight_layout()
    
    # Guardar en alta resolución
    plt.savefig(ruta_archivo, dpi=300, bbox_inches='tight')
    print(f"  ✓ Gráfico técnico guardado en: {ruta_archivo}")
    
    # Cierra la figura para evitar bloqueos en la ejecución
    plt.close()

def validar_coherencia_datos(D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg):
    errores, avisos = [], []
    if gamma_sat < gamma: errores.append("⚠️ γ_saturado < γ_natural")
    if D_w > D: avisos.append("ℹ️ Nivel freático bajo la base. Se asume D_w=D")
    if D_w < 0: errores.append("⚠️ Nivel freático negativo")
    if c < 0: errores.append("⚠️ Cohesión negativa")
    if phi_deg < 0 or phi_deg > 45: errores.append("⚠️ Ángulo φ fuera de rango [0°, 45°]")
    
    if errores:
        for e in errores: print(f"  {e}")
        return False, "Datos incoherentes"
    for a in avisos: print(f"  {a}")
    return True, "Datos validados"

def main():
    print("\n" + "="*80)
    print("   CÁLCULO DE CARGA ADMISIBLE PARA CIMENTACIONES SUPERFICIALES (CTE-DB-SE-C)")
    print("="*80)

    print("\n📋 PARÁMETROS DEL TERRENO:")
    D = obtener_float("  Profundidad de empotramiento (D, metros): ", 0.1)
    c = obtener_float("  Cohesión efectiva (c, kPa): ", 0)
    phi_deg = obtener_float("  Ángulo de rozamiento interno (φ, grados): ", 0, 45)
    gamma = obtener_float("  Peso específico natural (γ, kN/m³): ", 10, 30)
    gamma_sat = obtener_float("  Peso específico saturado (γ_sat, kN/m³): ", gamma, 30)
    D_w = obtener_float("  Profundidad del nivel freático (D_w, metros): ", 0)
    FS = obtener_float("  Factor de seguridad (FS, rango [1-10], default=3.0): ", 1, 10)
    alpha_deg = obtener_float("  Inclinación de la carga (α, grados, rango [0-90]): ", 0, 90)
    
    valido, msg = validar_coherencia_datos(D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg)
    if not valido: return
    
    considerar_sobrecarga = preguntar_sobrecarga()
    
    print("\n📏 DIMENSIONES DE LA ZAPATA:")
    B_inicio = obtener_float("  Valor inicial de B (ancho, metros): ", 0.1)
    B_fin = obtener_float("  Valor final de B (metros): ", B_inicio)
    B_paso = obtener_float("  Incremento de B (metros): ", 0.01)

    L_inicio = obtener_float("  Valor inicial de L (largo, metros): ", B_inicio)
    L_fin = obtener_float("  Valor final de L (metros): ", L_inicio)
    L_paso = obtener_float("  Incremento de L (metros): ", 0.01)

    combinaciones = generar_combinaciones_B_L(B_inicio, B_fin, B_paso, L_inicio, L_fin, L_paso)
    if not combinaciones:
        print("  ❌ No hay combinaciones válidas (B <= L)")
        return

    print(f"\n⚙️ Procesando {len(combinaciones)} combinaciones...")
    resultados = []
    for B, L in combinaciones:
        try:
            res = calcular_carga_admisible(B, L, D, c, phi_deg, gamma, gamma_sat, D_w, FS, alpha_deg, considerar_sobrecarga)
            resultados.append([B, L, *res])
        except Exception:
            pass

    # Mostrar preview en consola (primeros 5 y últimos 5 si hay muchos)
    print("\n📊 PREVIEW DE RESULTADOS:")
    print(f"{'B (m)':<8} {'L (m)':<8} {'q_ult (kPa)':<14} {'q_adm (kPa)':<14} {'Carga adm (kN)':<15}")
    print("-" * 65)
    for res in (resultados if len(resultados) <= 10 else resultados[:5] + [['...']*6] + resultados[-5:]):
        if res[0] == '...': print("  ... "*6)
        else: print(f"{res[0]:<8.2f} {res[1]:<8.2f} {res[2]:<14.2f} {res[4]:<14.2f} {res[5]:<15.2f}")

    exportar = input("\n¿Generar reporte completo (Excel + Gráfico)? (s/n): ").strip().lower()
    if exportar in ['s', 'si']:
        # 1. Crear nombre de carpeta con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_carpeta = f"Reporte_Cimentacion_{timestamp}"
        
        # 2. Crear el directorio físico
        os.makedirs(nombre_carpeta, exist_ok=True)
        print(f"\n📂 Creando directorio: ./{nombre_carpeta}/")

        parametros = {
            'D': D, 'c': c, 'phi': phi_deg, 'gamma': gamma, 'gamma_sat': gamma_sat,
            'D_w': D_w, 'FS': FS, 'alpha': alpha_deg, 'sobrecarga': 'Sí' if considerar_sobrecarga else 'No'
        }
        
        # 3. Exportar Excel dentro de la carpeta
        ruta_excel = os.path.join(nombre_carpeta, "tabla_resultados.xlsx")
        exportar_a_excel(resultados, ruta_excel, parametros)
        print(f"  ✓ Excel guardado en: {ruta_excel}")

        # 4. Generar y guardar el gráfico específico en la carpeta
        ruta_grafico = os.path.join(nombre_carpeta, "grafico_q_adm_vs_B.png")
        generar_grafico_especifico(resultados, parametros, ruta_grafico)
    
    print("\n✓ Proceso completado.\n")

if __name__ == "__main__":
    main()